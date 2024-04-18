# -*- coding: utf-8 -*-
# (c) Nano Nano Ltd 2019

import csv
import sys
import warnings
from typing import List, Optional, TextIO

import openpyxl
import xlrd
from colorama import Fore
from tqdm import tqdm, trange

from .config import config
from .constants import ERROR, FONT_COLOR_TX_DEST, FONT_COLOR_TX_HASH, FONT_COLOR_TX_SRC
from .exceptions import TransactionParserError
from .t_record import TransactionRecord
from .t_row import TransactionRow, TxRaw


class ImportRecords:
    def __init__(self) -> None:
        self.t_rows: List["TransactionRow"] = []
        self.success_cnt = 0
        self.failure_cnt = 0

    def import_excel_xlsx(self, filename: str) -> None:
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        workbook = openpyxl.load_workbook(filename=filename, read_only=True, data_only=True)
        print(f"{Fore.WHITE}Excel file: {Fore.YELLOW}{filename}")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            try:
                dimensions = worksheet.calculate_dimension()
            except ValueError:
                worksheet.reset_dimensions()
            else:
                if dimensions == "A1:A1" or dimensions.endswith("1048576"):
                    worksheet.reset_dimensions()

            if worksheet.title.startswith("--"):
                print(f"{Fore.GREEN}skipping '{worksheet.title}' worksheet")
                continue

            if config.debug:
                print(f"{Fore.CYAN}importing '{worksheet.title}' rows")

            for row_num, worksheet_row in enumerate(
                tqdm(
                    worksheet.rows,
                    total=worksheet.max_row,
                    unit=" row",
                    desc=f"{Fore.CYAN}importing '{worksheet.title}' rows{Fore.GREEN}",
                    disable=bool(config.debug or not sys.stdout.isatty()),
                )
            ):
                if row_num == 0:
                    # Skip headers
                    continue

                row = [self.convert_cell_xlsx(cell) for cell in worksheet_row]

                t_row = TransactionRow(
                    row[: len(TransactionRow.HEADER)], row_num + 1, filename, worksheet.title
                )

                try:
                    t_row.parse()
                except TransactionParserError as e:
                    t_row.failure = e
                else:
                    t_row.tx_raw = self.get_tx_raw_xlsx(worksheet_row)

                if config.debug or t_row.failure:
                    tqdm.write(f"{Fore.YELLOW}import: {t_row}")

                if t_row.failure:
                    tqdm.write(f"{ERROR} {t_row.failure}")

                self.t_rows.append(t_row)
                self.update_cnts(t_row)

        workbook.close()
        del workbook

    def import_excel_xls(self, filename: str) -> None:
        workbook = xlrd.open_workbook(filename)
        print(f"{Fore.WHITE}Excel file: {Fore.YELLOW}{filename}")

        for worksheet in workbook.sheets():
            if worksheet.name.startswith("--"):
                print(f"{Fore.GREEN}skipping '{worksheet.name}' worksheet")
                continue

            if config.debug:
                print(f"{Fore.CYAN}importing '{worksheet.name}' rows")

            for row_num in trange(
                0,
                worksheet.nrows,
                unit=" row",
                desc=f"{Fore.CYAN}importing '{worksheet.name}' rows{Fore.GREEN}",
                disable=bool(config.debug or not sys.stdout.isatty()),
            ):
                if row_num == 0:
                    # Skip headers
                    continue

                row = [
                    self.convert_cell_xls(worksheet.cell(row_num, cell_num), workbook)
                    for cell_num in range(0, worksheet.ncols)
                ]

                t_row = TransactionRow(
                    row[: len(TransactionRow.HEADER)], row_num + 1, filename, worksheet.name
                )

                try:
                    t_row.parse()
                except TransactionParserError as e:
                    t_row.failure = e

                if config.debug or t_row.failure:
                    tqdm.write(f"{Fore.YELLOW}import: {t_row}")

                if t_row.failure:
                    tqdm.write(f"{ERROR} {t_row.failure}")

                self.t_rows.append(t_row)
                self.update_cnts(t_row)

        workbook.release_resources()
        del workbook

    @staticmethod
    def convert_cell_xlsx(cell: openpyxl.cell.cell.Cell) -> str:
        if cell.value is None:
            return ""
        return str(cell.value)

    def get_tx_raw_xlsx(self, worksheet_row: List[openpyxl.cell.cell.Cell]) -> Optional["TxRaw"]:
        tx_hash = tx_src = tx_dest = ""

        for cell in worksheet_row[len(TransactionRow.HEADER) :]:
            if cell.value and cell.font.color.type == "rgb":
                if cell.font.color.rgb == f"FF{FONT_COLOR_TX_HASH}":
                    tx_hash = str(cell.value)
                elif cell.font.color.rgb == f"FF{FONT_COLOR_TX_SRC}":
                    tx_src = str(cell.value)
                elif cell.font.color.rgb == f"FF{FONT_COLOR_TX_DEST}":
                    tx_dest = str(cell.value)

        if any((tx_hash, tx_src, tx_dest)):
            return TxRaw(tx_hash, tx_src, tx_dest)
        return None

    @staticmethod
    def convert_cell_xls(cell: xlrd.sheet.Cell, workbook: xlrd.Book) -> str:
        if cell.ctype == xlrd.XL_CELL_DATE:
            datetime = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
            if datetime.microsecond:
                value = f"{datetime:%Y-%m-%dT%H:%M:%S.%f}"
            else:
                value = f"{datetime:%Y-%m-%dT%H:%M:%S}"
        elif cell.ctype in (
            xlrd.XL_CELL_NUMBER,
            xlrd.XL_CELL_BOOLEAN,
            xlrd.XL_CELL_ERROR,
        ):
            # repr is required to ensure no precision is lost
            value = repr(cell.value)
        else:
            value = str(cell.value)

        return value

    def import_csv(self, import_file: TextIO, filename: Optional[str] = None) -> None:
        print(f"{Fore.WHITE}CSV file: {Fore.YELLOW}{import_file.name}")
        if config.debug:
            print(f"{Fore.CYAN}importing rows")

        reader = csv.reader(import_file)

        for row in tqdm(
            reader,
            unit=" row",
            desc=f"{Fore.CYAN}importing{Fore.GREEN}",
            disable=bool(config.debug or not sys.stdout.isatty()),
        ):
            if reader.line_num == 1:
                # Skip headers
                continue

            t_row = TransactionRow(row[: len(TransactionRow.HEADER)], reader.line_num, filename)
            try:
                t_row.parse()
            except TransactionParserError as e:
                t_row.failure = e

            if config.debug or t_row.failure:
                tqdm.write(f"{Fore.YELLOW}import: {t_row}")

            if t_row.failure:
                tqdm.write(f"{ERROR} {t_row.failure}")

            self.t_rows.append(t_row)
            self.update_cnts(t_row)

    def update_cnts(self, t_row: "TransactionRow") -> None:
        if t_row.failure is not None:
            self.failure_cnt += 1
        elif t_row.t_record is not None:
            self.success_cnt += 1

    def get_records(self) -> List[TransactionRecord]:
        transaction_records = [t_row.t_record for t_row in self.t_rows if t_row.t_record]

        transaction_records.sort()
        for t_record in transaction_records:
            t_record.set_tid()

        if config.debug:
            for t_row in self.t_rows:
                print(f"{Fore.YELLOW}import: {t_row}")

        return transaction_records
